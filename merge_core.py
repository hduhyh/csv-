"""CSV/Excel merging engine used by the GUI and command-line entry point."""

from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Iterator, List, Optional, Sequence


ProgressCallback = Callable[[str], None]
SUPPORTED_INPUT_SUFFIXES = {".csv", ".xlsx", ".xlsm"}


@dataclass
class FileInfo:
    path: Path
    encoding: Optional[str]
    headers: List[str]
    row_count: int
    short_row_count: int = 0
    long_row_count: int = 0


@dataclass
class MergeResult:
    output_path: Path
    warning_path: Optional[Path]
    file_count: int
    row_count: int
    columns: List[str]
    warnings: List[Dict[str, str]] = field(default_factory=list)


def detect_encoding(file_path: Path) -> str:
    """Detect common encodings and validate the whole file in chunks."""

    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with file_path.open("r", encoding=encoding, newline="") as handle:
                while handle.read(1024 * 1024):
                    pass
            return encoding
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法识别文件编码：{file_path}")


def make_unique_headers(headers: Iterable[object]) -> List[str]:
    """Clean headers and make duplicate/blank names safe and deterministic."""

    result: List[str] = []
    counters: Dict[str, int] = {}
    for index, value in enumerate(headers, start=1):
        header = "" if value is None else str(value)
        header = header.replace("\ufeff", "").strip()
        if not header:
            header = f"__blank_col_{index}"

        count = counters.get(header, 0) + 1
        counters[header] = count
        result.append(header if count == 1 else f"{header}__dup{count}")
    return result


def _read_csv_template(template_path: Path) -> List[str]:
    encoding = detect_encoding(template_path)
    with template_path.open("r", encoding=encoding, newline="") as handle:
        first_row = next(csv.reader(handle), None)
    if not first_row:
        raise ValueError(f"列名模板没有可读取的第一行：{template_path}")
    return make_unique_headers(first_row)


def _read_excel_template(template_path: Path) -> List[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 模板需要安装 openpyxl。") from exc

    workbook = load_workbook(template_path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    finally:
        workbook.close()

    if not first_row or not any(value is not None and str(value).strip() for value in first_row):
        raise ValueError(f"列名模板没有可读取的第一行：{template_path}")

    values = list(first_row)
    while values and (values[-1] is None or str(values[-1]).strip() == ""):
        values.pop()
    return make_unique_headers(values)


def read_template_headers(template_path: Path) -> List[str]:
    suffix = template_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_template(template_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel_template(template_path)
    raise ValueError("列名模板仅支持 CSV、XLSX 或 XLSM 文件。")


def _trim_trailing_empty_cells(row: Iterable[Any]) -> List[Any]:
    """Remove Excel's trailing empty cells without changing interior blanks."""

    values = list(row)
    while values and values[-1] is None:
        values.pop()
    return values


def _excel_row_has_data(row: Sequence[Any]) -> bool:
    return any(value is not None and value != "" for value in row)


def _scan_csv(file_path: Path) -> FileInfo:
    encoding = detect_encoding(file_path)
    with file_path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle)
        raw_headers = next(reader, None)
        if raw_headers is None:
            return FileInfo(file_path, encoding, [], 0)

        rows = 0
        max_columns = len(raw_headers)
        row_length_counts: Counter[int] = Counter()
        for row in reader:
            rows += 1
            row_length = len(row)
            row_length_counts[row_length] += 1
            max_columns = max(max_columns, row_length)

    expanded_headers = list(raw_headers)
    if len(expanded_headers) < max_columns:
        expanded_headers.extend(
            f"__extra_col_{index}"
            for index in range(len(expanded_headers) + 1, max_columns + 1)
        )

    short_rows = sum(
        count for length, count in row_length_counts.items() if length < max_columns
    )
    long_rows = sum(
        count for length, count in row_length_counts.items() if length > len(raw_headers)
    )
    return FileInfo(
        path=file_path,
        encoding=encoding,
        headers=make_unique_headers(expanded_headers),
        row_count=rows,
        short_row_count=short_rows,
        long_row_count=long_rows,
    )


def _scan_excel(file_path: Path) -> FileInfo:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 输入文件需要安装 openpyxl。") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            return FileInfo(file_path, None, [], 0)

        sheet = workbook.worksheets[0]
        reader = sheet.iter_rows(values_only=True)
        first_row = next(reader, None)
        if first_row is None:
            return FileInfo(file_path, None, [], 0)

        raw_headers = _trim_trailing_empty_cells(first_row)
        rows = 0
        max_columns = len(raw_headers)
        row_length_counts: Counter[int] = Counter()
        for excel_row in reader:
            row = _trim_trailing_empty_cells(excel_row)
            if not _excel_row_has_data(row):
                continue
            rows += 1
            row_length = len(row)
            row_length_counts[row_length] += 1
            max_columns = max(max_columns, row_length)
    finally:
        workbook.close()

    expanded_headers = list(raw_headers)
    if len(expanded_headers) < max_columns:
        expanded_headers.extend(
            f"__extra_col_{index}"
            for index in range(len(expanded_headers) + 1, max_columns + 1)
        )

    short_rows = sum(
        count for length, count in row_length_counts.items() if length < max_columns
    )
    long_rows = sum(
        count for length, count in row_length_counts.items() if length > len(raw_headers)
    )
    return FileInfo(
        path=file_path,
        encoding=None,
        headers=make_unique_headers(expanded_headers),
        row_count=rows,
        short_row_count=short_rows,
        long_row_count=long_rows,
    )


def _scan_input_file(file_path: Path) -> FileInfo:
    if file_path.suffix.lower() == ".csv":
        return _scan_csv(file_path)
    return _scan_excel(file_path)


def _iter_input_rows(info: FileInfo) -> Iterator[List[Any]]:
    if info.path.suffix.lower() == ".csv":
        if not info.encoding:
            raise RuntimeError(f"CSV 文件缺少编码信息：{info.path}")
        with info.path.open("r", encoding=info.encoding, newline="") as input_handle:
            reader = csv.reader(input_handle)
            next(reader, None)
            for row in reader:
                yield row
        return

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 输入文件需要安装 openpyxl。") from exc

    workbook = load_workbook(info.path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        reader = sheet.iter_rows(values_only=True)
        next(reader, None)
        for row in reader:
            values = _trim_trailing_empty_cells(row)
            if _excel_row_has_data(values):
                yield values
    finally:
        workbook.close()


def _append_warning(
    warnings: List[Dict[str, str]], file_name: str, warning: str
) -> None:
    warnings.append({"file": file_name, "warning": warning})


def _write_warning_report(
    output_folder: Path, warnings: Sequence[Dict[str, str]]
) -> Optional[Path]:
    if not warnings:
        return None

    warning_path = output_folder / "csv_column_warning_report.csv"
    with warning_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["file", "warning"])
        writer.writeheader()
        writer.writerows(warnings)
    return warning_path


def merge_csv_files(
    input_folder: str,
    output_folder: Optional[str] = None,
    template_path: Optional[str] = None,
    output_filename: str = "merged_output.csv",
    add_source_column: bool = True,
    progress_callback: Optional[ProgressCallback] = None,
) -> MergeResult:
    """Merge top-level CSV/XLSX/XLSM files into one UTF-8 BOM CSV.

    If ``template_path`` is supplied, its first row defines the output columns
    and their order. Otherwise the union of source headers is used.
    """

    source_folder = Path(input_folder).expanduser()
    destination_folder = Path(output_folder or input_folder).expanduser()
    template = Path(template_path).expanduser() if template_path else None

    if not source_folder.is_dir():
        raise FileNotFoundError(f"输入文件夹不存在：{source_folder}")
    if template and not template.is_file():
        raise FileNotFoundError(f"列名模板不存在：{template}")

    output_filename = output_filename.strip()
    if not output_filename:
        raise ValueError("输出文件名不能为空。")
    if Path(output_filename).name != output_filename:
        raise ValueError("输出文件名不能包含文件夹路径。")
    if any(character in output_filename for character in '<>:"/\\|?*'):
        raise ValueError("输出文件名包含 Windows 不支持的字符。")
    if output_filename.endswith((" ", ".")):
        raise ValueError("输出文件名不能以空格或句点结尾。")
    if Path(output_filename).suffix.lower() != ".csv":
        output_filename += ".csv"
    reserved_names = {"CON", "PRN", "AUX", "NUL"}
    reserved_names.update(f"COM{index}" for index in range(1, 10))
    reserved_names.update(f"LPT{index}" for index in range(1, 10))
    if Path(output_filename).stem.upper() in reserved_names:
        raise ValueError("该输出文件名是 Windows 保留名称，请换一个名称。")

    destination_folder.mkdir(parents=True, exist_ok=True)
    output_path = destination_folder / output_filename
    resolved_output = output_path.resolve()
    resolved_template = template.resolve() if template else None
    if resolved_template == resolved_output:
        raise ValueError("列名模板不能与输出文件使用同一个路径。")

    input_files = []
    candidates = sorted(source_folder.iterdir(), key=lambda item: item.name.lower())
    for path in candidates:
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES:
            continue
        if path.name.startswith("~$") and path.suffix.lower() in {".xlsx", ".xlsm"}:
            continue
        resolved = path.resolve()
        if (
            resolved == resolved_output
            or resolved == resolved_template
            or path.name.lower() == output_filename.lower()
        ):
            continue
        if path.name.lower() == "csv_column_warning_report.csv":
            continue
        input_files.append(path)

    if not input_files:
        raise FileNotFoundError(
            f"文件夹内没有找到可合并的 CSV 或 Excel 文件：{source_folder}"
        )

    def report(message: str) -> None:
        if progress_callback:
            progress_callback(message)

    report(
        f"找到 {len(input_files)} 个 CSV/Excel 文件，正在检查工作表、列名和编码……"
    )
    file_infos: List[FileInfo] = []
    warnings: List[Dict[str, str]] = []
    union_columns: List[str] = []

    for index, input_file in enumerate(input_files, start=1):
        report(f"检查 ({index}/{len(input_files)})：{input_file.name}")
        info = _scan_input_file(input_file)
        if not info.headers:
            _append_warning(warnings, input_file.name, "空文件或首个工作表为空，已跳过")
            continue
        file_infos.append(info)
        for header in info.headers:
            if header not in union_columns:
                union_columns.append(header)
        if info.long_row_count:
            _append_warning(
                warnings,
                input_file.name,
                f"有 {info.long_row_count} 行的数据列多于原始表头，已自动补充占位列名",
            )
        if info.short_row_count:
            _append_warning(
                warnings,
                input_file.name,
                f"有 {info.short_row_count} 行列数不足，已在合并时补空值",
            )

    if not file_infos:
        raise ValueError("没有成功读取到任何有效的 CSV 或 Excel 文件。")

    if template:
        output_columns = read_template_headers(template)
        report(f"已读取模板，共 {len(output_columns)} 个目标列。")
        for info in file_infos:
            missing = [column for column in output_columns if column not in info.headers]
            extra = [column for column in info.headers if column not in output_columns]
            if missing:
                _append_warning(
                    warnings,
                    info.path.name,
                    f"缺少模板中的 {len(missing)} 列，已补空值：{', '.join(missing[:10])}"
                    + ("……" if len(missing) > 10 else ""),
                )
            if extra:
                _append_warning(
                    warnings,
                    info.path.name,
                    f"有 {len(extra)} 列不在模板中，未写入：{', '.join(extra[:10])}"
                    + ("……" if len(extra) > 10 else ""),
                )
    else:
        output_columns = union_columns

    source_column = "Source_File"
    if add_source_column and source_column not in output_columns:
        output_columns = list(output_columns) + [source_column]

    report("开始写入合并结果……")
    total_rows = 0
    temporary_output = output_path.with_suffix(output_path.suffix + ".tmp")
    try:
        with temporary_output.open("w", encoding="utf-8-sig", newline="") as output_handle:
            writer = csv.writer(output_handle, lineterminator="\n")
            writer.writerow(output_columns)

            for index, info in enumerate(file_infos, start=1):
                report(f"合并 ({index}/{len(file_infos)})：{info.path.name}")
                for row in _iter_input_rows(info):
                    values = row + [""] * max(0, len(info.headers) - len(row))
                    row_map = dict(zip(info.headers, values))
                    if add_source_column:
                        row_map[source_column] = info.path.name
                    writer.writerow([row_map.get(column, "") for column in output_columns])
                    total_rows += 1
        temporary_output.replace(output_path)
    except Exception:
        temporary_output.unlink(missing_ok=True)
        raise

    warning_path = _write_warning_report(destination_folder, warnings)
    report(f"完成：已合并 {len(file_infos)} 个文件、{total_rows} 行数据。")
    return MergeResult(
        output_path=output_path,
        warning_path=warning_path,
        file_count=len(file_infos),
        row_count=total_rows,
        columns=output_columns,
        warnings=warnings,
    )
