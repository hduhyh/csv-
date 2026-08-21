from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from merge_core import make_unique_headers, merge_csv_files, read_template_headers


def write_csv(path: Path, rows: list[list[str]], encoding: str = "utf-8") -> None:
    with path.open("w", encoding=encoding, newline="") as handle:
        csv.writer(handle).writerows(rows)


def read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def write_xlsx(path: Path, rows: list[list[object]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


class MergeCoreTests(unittest.TestCase):
    def test_merge_without_template_uses_union_and_source_file(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            write_csv(source / "a.csv", [["id", "name"], ["1", "甲"]])
            write_csv(source / "b.csv", [["id", "score"], ["2", "99"]])

            result = merge_csv_files(str(source), str(output))

            self.assertEqual(result.file_count, 2)
            self.assertEqual(result.row_count, 2)
            self.assertEqual(
                read_csv(result.output_path),
                [
                    ["id", "name", "score", "Source_File"],
                    ["1", "甲", "", "a.csv"],
                    ["2", "", "99", "b.csv"],
                ],
            )

    def test_csv_template_controls_columns_and_order(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            template = root / "template.csv"
            write_csv(source / "data.csv", [["id", "name", "extra"], ["1", "甲", "x"]])
            write_csv(template, [["name", "missing", "id"]])

            result = merge_csv_files(
                str(source),
                str(output),
                template_path=str(template),
                add_source_column=False,
            )

            self.assertEqual(
                read_csv(result.output_path),
                [["name", "missing", "id"], ["甲", "", "1"]],
            )
            self.assertIsNotNone(result.warning_path)
            warning_text = result.warning_path.read_text(encoding="utf-8-sig")
            self.assertIn("缺少模板中的 1 列", warning_text)
            self.assertIn("有 1 列不在模板中", warning_text)

    def test_csv_and_xlsx_inputs_are_merged_by_column_name(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            write_csv(source / "a.csv", [["id", "name"], ["1", "CSV数据"]])
            write_xlsx(
                source / "b.xlsx",
                [["name", "id", "score"], ["Excel数据", 2, 98.5]],
            )

            result = merge_csv_files(str(source), str(output))

            self.assertEqual(result.file_count, 2)
            self.assertEqual(result.row_count, 2)
            self.assertEqual(
                read_csv(result.output_path),
                [
                    ["id", "name", "score", "Source_File"],
                    ["1", "CSV数据", "", "a.csv"],
                    ["2", "Excel数据", "98.5", "b.xlsx"],
                ],
            )

    def test_xlsx_input_reads_only_first_worksheet(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook = Workbook()
            first_sheet = workbook.active
            first_sheet.title = "需要合并"
            first_sheet.append(["id", "value"])
            first_sheet.append([1, "第一页"])
            first_sheet.append([None, None])
            second_sheet = workbook.create_sheet("不应合并")
            second_sheet.append(["id", "value"])
            second_sheet.append([2, "第二页"])
            workbook.save(root / "book.xlsx")

            result = merge_csv_files(str(root), add_source_column=False)

            self.assertEqual(result.row_count, 1)
            self.assertEqual(
                read_csv(result.output_path),
                [["id", "value"], ["1", "第一页"]],
            )

    def test_excel_template_inside_input_folder_is_not_merged(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            write_csv(root / "data.csv", [["id", "name"], ["1", "测试"]])
            template = root / "template.xlsx"
            write_xlsx(template, [["name", "id"]])

            result = merge_csv_files(
                str(root),
                template_path=str(template),
                add_source_column=False,
            )

            self.assertEqual(result.file_count, 1)
            self.assertEqual(
                read_csv(result.output_path),
                [["name", "id"], ["测试", "1"]],
            )

    def test_gbk_input_and_existing_output_are_handled(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            write_csv(root / "中文.csv", [["编号", "名称"], ["1", "测试"]], "gbk")

            first = merge_csv_files(str(root), str(root), add_source_column=False)
            second = merge_csv_files(str(root), str(root), add_source_column=False)

            self.assertEqual(first.row_count, 1)
            self.assertEqual(second.row_count, 1)
            self.assertEqual(read_csv(second.output_path), [["编号", "名称"], ["1", "测试"]])

    def test_existing_output_name_is_skipped_when_output_folder_changes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "source"
            output = root / "output"
            source.mkdir()
            write_csv(source / "data.csv", [["id"], ["1"]])
            write_csv(source / "merged_output.csv", [["id"], ["old"], ["old"]])

            result = merge_csv_files(str(source), str(output), add_source_column=False)

            self.assertEqual(result.file_count, 1)
            self.assertEqual(result.row_count, 1)
            self.assertEqual(read_csv(result.output_path), [["id"], ["1"]])

    def test_extra_data_columns_are_preserved_without_template(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            write_csv(root / "data.csv", [["id"], ["1", "extra"], ["2"]])

            result = merge_csv_files(str(root), add_source_column=False)

            self.assertEqual(
                read_csv(result.output_path),
                [["id", "__extra_col_2"], ["1", "extra"], ["2", ""]],
            )
            self.assertIsNotNone(result.warning_path)

    def test_excel_template_reads_first_row(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as temporary_directory:
            template = Path(temporary_directory) / "template.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["name", "id", None])
            workbook.save(template)

            self.assertEqual(read_template_headers(template), ["name", "id"])

    def test_headers_are_cleaned_and_deduplicated(self) -> None:
        self.assertEqual(
            make_unique_headers([" id ", "", "id", None]),
            ["id", "__blank_col_2", "id__dup2", "__blank_col_4"],
        )

    def test_windows_invalid_output_name_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            write_csv(root / "data.csv", [["id"], ["1"]])

            with self.assertRaisesRegex(ValueError, "Windows"):
                merge_csv_files(str(root), output_filename="bad:name.csv")
            with self.assertRaisesRegex(ValueError, "保留名称"):
                merge_csv_files(str(root), output_filename="CON.csv")


if __name__ == "__main__":
    unittest.main()
